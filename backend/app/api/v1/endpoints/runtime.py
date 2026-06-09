import csv
import io
import json
import uuid
import logging
from datetime import datetime
from typing import List, Optional, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, update

from app.db.base import get_db
from app.models.runtime import (
    RuntimeDataCenter,
    RuntimeAsset,
    DataSourceImport,
    ApplicationIntent,
    SourceProposal,
    RuntimeAuditLog
)
from app.services.confidence_service import engine as confidence_engine
from app.services.drift_service import (
    run_drift_detection,
    run_drift_detection_all,
    compute_alignment_status,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/runtime-location", tags=["runtime-location"])

# ─── Utility Parsers ─────────────────────────────────────────────────────────

def detect_source_type(file_name: str) -> str:
    f = file_name.lower()
    if "ibmma" in f or "qmgr" in f or ("mq" in f and "mongo" not in f):
        return "ibm_mq"
    if "mongodb" in f or "mongo_info" in f or "mongo" in f:
        return "mongodb"
    if "oem" in f or "oracle" in f or "db_role" in f:
        return "oracle_oem"
    if "scom" in f or "replica_status" in f or "replicastatus" in f:
        return "scom"
    if "mssql" in f or "sqlserver" in f or "sql_server" in f:
        return "mssql"
    if "kafka" in f:
        return "kafka"
    if "avi" in f or "loadbalancer" in f or "load_balancer" in f or "gslb" in f or "virtual" in f:
        return "avi_loadbalancer"
    if "ocp" in f or "pod_info" in f or "openshift" in f:
        return "ocp"
    if "batch" in f or "batch_processing" in f or "jobs" in f:
        return "batch"
    if "appdynamics" in f or "appdynamic" in f or "node_inventory" in f:
        return "appdynamics"
    return "cmdb"

def resolve_dc_from_mq_hostname(hostname: str) -> Dict[str, str]:
    h = hostname.lower()
    if h.startswith("mq4uprdga") or "prdga" in h:
        return {"name": "DC Georgia Production", "short_name": "GA-PRD"}
    if h.startswith("mq4uprdma") or "prdma" in h:
        return {"name": "DC Maryland Production", "short_name": "MA-PRD"}
    if h.startswith("mq4uatga") or "uatga" in h:
        return {"name": "DC Georgia UAT", "short_name": "GA-UAT"}
    if h.startswith("mq4uatma") or "uatma" in h:
        return {"name": "DC Maryland UAT", "short_name": "MA-UAT"}
    if "prd" in h or "prod" in h:
        return {"name": "DC Production", "short_name": "PRD"}
    if "uat" in h:
        return {"name": "DC UAT", "short_name": "UAT"}
    return {"name": "DC Unknown (Inferred)", "short_name": "UNK"}

def resolve_dc_from_mongo_hostname(hostname: str) -> Dict[str, str]:
    h = hostname.lower()
    if "az" in h:
        import re
        match = re.search(r"az(\d+)", h)
        if match:
            return {"name": f"Azure Zone {match.group(1)}", "short_name": f"AZ{match.group(1)}"}
    if "prod" in h or "prd" in h:
        return {"name": "DC Production", "short_name": "PRD"}
    if "uat" in h:
        return {"name": "DC UAT", "short_name": "UAT"}
    return {"name": "DC Cloud (Inferred)", "short_name": "CLD"}

def resolve_dc_from_oracle_hostname(hostname: str) -> Dict[str, str]:
    h = hostname.lower()
    if "ibb1" in h:
        return {"name": "DC Birmingham IBB1", "short_name": "IBB1"}
    if "shv" in h:
        return {"name": "DC Shoreview", "short_name": "SHV"}
    if "uat" in h:
        return {"name": "DC UAT", "short_name": "UAT"}
    if "prod" in h or "prd" in h:
        return {"name": "DC Production", "short_name": "PRD"}
    return {"name": "DC Unknown (Inferred)", "short_name": "UNK"}

def resolve_dc_from_mssql_hostname(hostname: str) -> Dict[str, str]:
    h = hostname.lower()
    if "ibb1" in h:
        return {"name": "DC Birmingham IBB1", "short_name": "IBB1"}
    if "shv" in h:
        return {"name": "DC Shoreview", "short_name": "SHV"}
    return {"name": "DC Cloud (Inferred)", "short_name": "CLD"}

def resolve_dc_from_kafka_hostname(hostname: str) -> Dict[str, str]:
    h = hostname.lower()
    if "ibb1" in h:
        return {"name": "DC Birmingham IBB1", "short_name": "IBB1"}
    if "shv" in h:
        return {"name": "DC Shoreview", "short_name": "SHV"}
    return {"name": "DC Cloud (Inferred)", "short_name": "CLD"}

def resolve_dc_from_avi_hostname(hostname: str) -> Dict[str, str]:
    h = hostname.lower()
    if "ibb1" in h:
        return {"name": "DC Birmingham IBB1", "short_name": "IBB1"}
    if "shv" in h:
        return {"name": "DC Shoreview", "short_name": "SHV"}
    return {"name": "DC Cloud (Inferred)", "short_name": "CLD"}

async def get_or_create_dc(db: AsyncSession, dc_info: Dict[str, str]) -> RuntimeDataCenter:
    short_name = dc_info["short_name"]
    result = await db.execute(select(RuntimeDataCenter).where(RuntimeDataCenter.short_name == short_name))
    dc = result.scalar_one_or_none()
    if not dc:
        dc = RuntimeDataCenter(
            id=str(uuid.uuid4()),
            name=dc_info["name"],
            short_name=short_name,
            asset_count=0
        )
        db.add(dc)
        await db.flush()
    return dc

# ─── Core Conflict Detection ───────────────────────────────────────────────

def detect_conflicts(assets: List[RuntimeAsset]) -> List[Dict[str, Any]]:
    conflicts = []
    # Group by normalized host name
    by_host = {}
    for a in assets:
        if not a.host:
            continue
        key = a.host.lower().strip()
        if key not in by_host:
            by_host[key] = []
        by_host[key].append(a)

    for key, group in by_host.items():
        if len(group) < 2:
            continue
        sources = list(set(a.data_source for a in group))
        if len(sources) < 2:
            continue

        roles = list(set(a.latest_replication_role for a in group if a.latest_replication_role))
        states = list(set(a.latest_operational_state for a in group if a.latest_operational_state))
        dcs = list(set(a.data_center_short for a in group if a.data_center_short))

        a = group[0]
        b = next((g for g in group if g.data_source != a.data_source), group[1])

        if len(roles) > 1:
            conflicts.append({
                "asset_name": a.name,
                "source_a": {"name": a.data_source, "says": a.latest_replication_role or "UNKNOWN"},
                "source_b": {"name": b.data_source, "says": b.latest_replication_role or "UNKNOWN"},
                "last_checked": datetime.utcnow().isoformat() + "Z"
            })
        elif len(states) > 1:
            conflicts.append({
                "asset_name": a.name,
                "source_a": {"name": a.data_source, "says": f"state={a.latest_operational_state}"},
                "source_b": {"name": b.data_source, "says": f"state={b.latest_operational_state}"},
                "last_checked": datetime.utcnow().isoformat() + "Z"
            })
        elif len(dcs) > 1:
            conflicts.append({
                "asset_name": a.name,
                "source_a": {"name": a.data_source, "says": f"dc={a.data_center_short}"},
                "source_b": {"name": b.data_source, "says": f"dc={b.data_center_short}"},
                "last_checked": datetime.utcnow().isoformat() + "Z"
            })

        # Internal MongoDB conflict check
        if a.metadata_json and a.metadata_json.get("internal_conflict"):
            conflict_parts = a.metadata_json["internal_conflict"].split("vs")
            conflicts.append({
                "asset_name": a.name,
                "source_a": {"name": "mongodb_text", "says": conflict_parts[0].replace("text=", "").strip()},
                "source_b": {"name": "mongodb_int", "says": conflict_parts[1].replace("int=", "").strip()},
                "last_checked": datetime.utcnow().isoformat() + "Z"
            })

    return conflicts

# ─── API Endpoints ───────────────────────────────────────────────────────────

@router.get("/applications", response_model=List[Dict[str, Any]])
async def get_applications(db: AsyncSession = Depends(get_db)):
    # Pull all assets
    result = await db.execute(select(RuntimeAsset))
    assets = result.scalars().all()

    # Query intents
    intent_res = await db.execute(select(ApplicationIntent))
    intents = {i.application_id: i for i in intent_res.scalars().all()}

    # Group by app_id (CMDB application_id or fallback by tech stack infra)
    app_groups = {}
    for a in assets:
        app_id = "INFRASTRUCTURE"
        app_name = "Infrastructure Services"
        if a.metadata_json and a.metadata_json.get("application_id"):
            app_id = a.metadata_json["application_id"]
            app_name = a.metadata_json.get("application_name", app_id)
        elif a.data_source == "ibm_mq":
            app_id = "MQ_INFRA"
            app_name = "IBM MQ Shared Tier"
        elif a.data_source == "mongodb":
            app_id = "MONGO_INFRA"
            app_name = "MongoDB DB Tier"
        elif a.data_source == "oracle_oem":
            app_id = "ORACLE_INFRA"
            app_name = "Oracle Core databases"

        key = (app_id, a.environment)
        if key not in app_groups:
            app_groups[key] = {
                "application_id": app_id,
                "application_name": app_name,
                "environment": a.environment,
                "data_centers": set(),
                "tech_stacks": set(),
                "overall_confidence": [],
                "_assets": [],
                "asset_count": 0,
                "stale_source_count": 0,
                "last_updated": a.last_seen_at
            }

        group = app_groups[key]
        if a.data_center_short:
            group["data_centers"].add(a.data_center_short)
        group["tech_stacks"].add(a.tech_stack)
        group["overall_confidence"].append(a.latest_confidence_level)
        group["_assets"].append(a)
        group["asset_count"] += 1
        if a.last_seen_at > group["last_updated"]:
            group["last_updated"] = a.last_seen_at

    # Load projects, teams, lobs for metadata linking
    from app.models.project import Project
    from app.models.team import Team
    from app.models.lob import Lob

    proj_res = await db.execute(select(Project))
    projects_map = {p.id: p for p in proj_res.scalars().all()}

    team_res = await db.execute(select(Team))
    teams_map = {t.id: t for t in team_res.scalars().all()}

    lob_res = await db.execute(select(Lob))
    lobs_map = {l.id: l for l in lob_res.scalars().all()}

    summaries = []
    for key, data in app_groups.items():
        app_id = data["application_id"]
        group_assets = data.get("_assets", [])

        # Use confidence engine for deterministic scoring
        if group_assets:
            conf_result = confidence_engine.score_application(group_assets)
            conf_label = conf_result.level
            conf_score = conf_result.score
            # Map to legacy 1-4 numeric level
            level_map = {"HIGH": 4, "MEDIUM": 3, "LOW": 2, "CONFLICT": 2, "UNKNOWN": 1}
            conf_numeric = level_map.get(conf_label, 3)
        else:
            conf_numeric = int(min(data["overall_confidence"])) if data["overall_confidence"] else 3
            conf_label = {4: "HIGH", 3: "MEDIUM", 2: "LOW", 1: "UNKNOWN"}.get(conf_numeric, "MEDIUM")
            conf_score = {4: 90, 3: 65, 2: 45, 1: 0}.get(conf_numeric, 65)

        intent = intents.get(app_id)
        alignment_status = intent.alignment_status if intent else "UNKNOWN"

        # Linked metadata
        project_id = intent.project_id if intent else None
        project_name = None
        team_id = None
        team_name = None
        lob_id = None
        lob_name = None

        if project_id and project_id in projects_map:
            p = projects_map[project_id]
            project_name = p.name
            team_id = p.team_id
            lob_id = p.lob_id
            if team_id and team_id in teams_map:
                team_name = teams_map[team_id].name
            if lob_id and lob_id in lobs_map:
                lob_name = lobs_map[lob_id].name

        summaries.append({
            "application_id": app_id,
            "application_name": data["application_name"],
            "environment": data["environment"],
            "data_centers": list(data["data_centers"]),
            "tech_stacks": list(data["tech_stacks"]),
            "overall_confidence": conf_numeric,
            "confidence_label": conf_label,
            "confidence_score": conf_score,
            "component_count": 1 if app_id.endswith("INFRA") else 3,
            "asset_count": data["asset_count"],
            "stale_source_count": 0,
            "alignment_status": alignment_status,
            "last_updated": data["last_updated"].isoformat() + "Z",
            "project_id": project_id,
            "project_name": project_name,
            "team_id": team_id,
            "team_name": team_name,
            "lob_id": lob_id,
            "lob_name": lob_name,
        })

    return summaries

@router.get("/applications/{app_id}", response_model=Dict[str, Any])
async def get_application_detail(app_id: str, environment: str = "PRODUCTION", db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(RuntimeAsset))
    assets = result.scalars().all()

    # Filter assets belonging to this app
    app_assets = []
    for a in assets:
        is_match = False
        if a.metadata_json and a.metadata_json.get("application_id") == app_id:
            is_match = True
        elif app_id == "MQ_INFRA" and a.data_source == "ibm_mq":
            is_match = True
        elif app_id == "MONGO_INFRA" and a.data_source == "mongodb":
            is_match = True
        elif app_id == "ORACLE_INFRA" and a.data_source == "oracle_oem":
            is_match = True
        elif app_id == "INFRASTRUCTURE" and not (a.metadata_json and a.metadata_json.get("application_id")):
            is_match = True
        
        if is_match and a.environment == environment:
            app_assets.append(a)

    # Resolve LOB, Team, Project Hierarchy and Telemetry deep links
    intent_res = await db.execute(select(ApplicationIntent).where(ApplicationIntent.application_id == app_id))
    intent = intent_res.scalar_one_or_none()

    project_id = intent.project_id if intent else None
    project_name = None
    team_id = None
    team_name = None
    lob_id = None
    lob_name = None
    telemetry_links = []

    if project_id:
        from app.models.project import Project
        from app.models.team import Team
        from app.models.lob import Lob
        from app.models.project_connector import ProjectConnector
        from app.models.connector_catalog import ConnectorCatalogEntry
        import json

        proj_res = await db.execute(select(Project).where(Project.id == project_id))
        proj = proj_res.scalar_one_or_none()
        if proj:
            project_name = proj.name
            team_id = proj.team_id
            lob_id = proj.lob_id

            if team_id:
                t_res = await db.execute(select(Team).where(Team.id == team_id))
                t_val = t_res.scalar_one_or_none()
                if t_val:
                    team_name = t_val.name
            
            if lob_id:
                l_res = await db.execute(select(Lob).where(Lob.id == lob_id))
                l_val = l_res.scalar_one_or_none()
                if l_val:
                    lob_name = l_val.name

            # Fetch project connectors
            pc_res = await db.execute(
                select(ProjectConnector, ConnectorCatalogEntry)
                .join(ConnectorCatalogEntry, ProjectConnector.catalog_entry_id == ConnectorCatalogEntry.id)
                .where(ProjectConnector.project_id == project_id, ProjectConnector.is_enabled == True)
            )
            for pc, catalog in pc_res.all():
                url = None
                try:
                    config_data = json.loads(pc.config) if pc.config else {}
                    url = config_data.get("api_url") or config_data.get("base_url") or config_data.get("url") or config_data.get("connection_string")
                except Exception:
                    pass
                
                if not url:
                    url = f"https://{catalog.slug}.corp.internal/apps/{app_id.lower()}"

                telemetry_links.append({
                    "id": pc.id,
                    "name": pc.name,
                    "slug": catalog.slug,
                    "category": catalog.category,
                    "url": url,
                    "color": catalog.color or "#0A84FF",
                    "icon": catalog.icon or "link"
                })

    if not app_assets:
        return {
            "application_id": app_id,
            "application_name": app_id.replace("_", " ").title(),
            "environment": environment,
            "overall_confidence": 3,
            "components": [],
            "data_sources": [],
            "conflicts": [],
            "project_id": project_id,
            "project_name": project_name,
            "team_id": team_id,
            "team_name": team_name,
            "lob_id": lob_id,
            "lob_name": lob_name,
            "telemetry_links": telemetry_links,
        }

    # Group into components based on tech stack/asset type
    stack_groups = {}
    for a in app_assets:
        stack = a.tech_stack
        if stack not in stack_groups:
            stack_groups[stack] = []
        stack_groups[stack].append(a)

    components = []
    for stack, stack_assets in stack_groups.items():
        comp_type = "COMPUTE"
        if stack in ["oracle", "mongodb", "mssql"]:
            comp_type = "DATABASE"
        elif stack in ["ibm_mq", "kafka"]:
            comp_type = "MESSAGING"

        # Map to component model
        mapped_assets = []
        for sa in stack_assets:
            mapped_assets.append({
                "id": sa.id,
                "name": sa.name,
                "asset_type": sa.asset_type,
                "tech_stack": sa.tech_stack,
                "environment": sa.environment,
                "host": sa.host,
                "port": sa.port,
                "platform": sa.platform,
                "data_center": {
                    "id": f"dc-{sa.data_center_short.lower()}" if sa.data_center_short else "dc-unknown",
                    "name": sa.data_center_short or "Unknown DC",
                    "short_name": sa.data_center_short or "UNK",
                    "asset_count": 1
                },
                "latest_confidence_level": sa.latest_confidence_level,
                "latest_operational_state": sa.latest_operational_state,
                "latest_replication_role": sa.latest_replication_role,
                "write_authority": sa.write_authority,
                "last_seen_at": sa.last_seen_at.isoformat() + "Z",
                "is_deterministic": sa.is_deterministic,
                "data_source": sa.data_source,
                "metadata": sa.metadata_json
            })

        components.append({
            "id": f"comp-{app_id}-{stack}",
            "application_id": app_id,
            "application_name": app_assets[0].metadata_json.get("application_name", app_id) if app_assets[0].metadata_json else app_id,
            "component_name": f"{stack.replace('_', ' ').upper()} Layer",
            "component_type": comp_type,
            "tech_stack": stack,
            "assets": mapped_assets
        })

    conflicts = detect_conflicts(app_assets)

    # Calculate overall confidence using the engine
    conf_result = confidence_engine.score_application(app_assets)
    conf_label = conf_result.level
    conf_score = conf_result.score
    level_map = {"HIGH": 4, "MEDIUM": 3, "LOW": 2, "CONFLICT": 2, "UNKNOWN": 1}
    min_conf = level_map.get(conf_label, 3)

    # Built dynamic data sources info
    sources_set = list(set(a.data_source for a in app_assets))
    data_sources = []
    for s in sources_set:
        source_assets = [a for a in app_assets if a.data_source == s]
        fresh = True
        confs = [a.latest_confidence_level for a in source_assets]
        data_sources.append({
            "source_name": s,
            "display_name": s.replace("_", " ").title(),
            "record_count": len(source_assets),
            "confidence_level": min(confs) if confs else 3,
            "last_imported": max(a.last_seen_at for a in source_assets).isoformat() + "Z",
            "is_fresh": fresh
        })

    return {
        "application_id": app_id,
        "application_name": app_assets[0].metadata_json.get("application_name", app_id) if app_assets[0].metadata_json else app_id,
        "environment": environment,
        "overall_confidence": min_conf,
        "confidence_label": conf_label,
        "confidence_score": conf_score,
        "components": components,
        "data_sources": data_sources,
        "conflicts": conflicts,
        "project_id": project_id,
        "project_name": project_name,
        "team_id": team_id,
        "team_name": team_name,
        "lob_id": lob_id,
        "lob_name": lob_name,
        "telemetry_links": telemetry_links,
    }

@router.get("/datacenters", response_model=List[Dict[str, Any]])
async def get_datacenters(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(RuntimeDataCenter))
    dcs = result.scalars().all()
    
    # Refresh asset counts
    for dc in dcs:
        asset_res = await db.execute(select(RuntimeAsset).where(RuntimeAsset.data_center_short == dc.short_name))
        dc.asset_count = len(asset_res.scalars().all())
    
    return [
        {
            "id": dc.id,
            "name": dc.name,
            "short_name": dc.short_name,
            "region": dc.region,
            "zone": dc.zone,
            "asset_count": dc.asset_count
        } for dc in dcs
    ]
@router.get("/imports", response_model=List[Dict[str, Any]])
async def get_imports(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(DataSourceImport).order_by(DataSourceImport.imported_at.desc()))
    imports = result.scalars().all()
    return [
        {
            "id": imp.id,
            "source_name": imp.source_name,
            "file_name": imp.file_name,
            "imported_at": imp.imported_at.isoformat() + "Z",
            "record_count": imp.record_count,
            "status": imp.status,
            "errors": imp.errors
        } for imp in imports
    ]

async def parse_and_insert_csv(
    filename: str,
    content: str,
    source_type: Optional[str],
    db: AsyncSession
) -> Dict[str, Any]:
    import io
    import csv
    import uuid
    import re
    from datetime import datetime
    
    source = source_type or detect_source_type(filename)
    errors = []
    assets_created = 0
    
    f = io.StringIO(content)
    reader = csv.DictReader(f)
    
    try:
        if source == "ibm_mq":
            for row in reader:
                hostname = row.get("hostname") or row.get("HOSTNAME")
                qmgr = row.get("qmgr") or row.get("QMGR") or hostname
                env = (row.get("env") or row.get("ENV") or "UAT").upper()
                platform = row.get("platform") or row.get("PLATFORM") or "UNIX"
                port = int(row.get("port") or row.get("PORT") or "1414")
                cluster = row.get("cluster") or ""
                exported_qmgr = row.get("exported_qmgr") or ""
                mq_namespace = row.get("mq_namespace") or ""

                if not hostname:
                    continue
                
                dc_info = resolve_dc_from_mq_hostname(hostname)
                dc = await get_or_create_dc(db, dc_info)
                
                asset_env = "PRODUCTION" if env in ["PRODUCTION", "PROD"] else "DR" if env == "DR" else "UAT"
                conf = 4 if cluster else 3

                # Extract application_id
                app_id = "MQ_INFRA"
                app_name = "IBM MQ Shared Tier"
                ref_name = qmgr or exported_qmgr or ""
                if ref_name:
                    match = re.search(r'[A-Z]{3,}', ref_name)
                    if match:
                        app_id = match.group(0)
                        app_name = f"{app_id} Messaging Tier"

                asset = RuntimeAsset(
                    id=str(uuid.uuid4()),
                    name=qmgr,
                    asset_type="MQ_QMGR",
                    tech_stack="ibm_mq",
                    environment=asset_env,
                    host=hostname,
                    port=port,
                    platform=platform,
                    data_center_short=dc.short_name,
                    latest_confidence_level=conf,
                    latest_operational_state="ACTIVE",
                    latest_replication_role="NONE",
                    write_authority=True,
                    is_deterministic=True,
                    data_source="ibm_mq",
                    metadata_json={
                        "cluster": cluster,
                        "exported_qmgr": exported_qmgr,
                        "mq_namespace": mq_namespace,
                        "cluster_role": "CLUSTER_MEMBER" if cluster else "STANDALONE",
                        "application_id": app_id,
                        "application_name": app_name
                    }
                )
                db.add(asset)
                assets_created += 1

        elif source == "mongodb":
            for row in reader:
                hostname = row.get("hostname") or row.get("HOSTNAME")
                replica_state = (row.get("replica_state_name") or row.get("REPLICA_STATE_NAME") or "SECONDARY").upper()
                rs_nm = row.get("rs_nm") or row.get("RS_NM") or ""
                env = (row.get("env") or row.get("ENV") or "UAT").upper()
                cluster = row.get("cluster") or row.get("CLUSTER") or ""
                role = (row.get("cl_role") or row.get("role") or row.get("ROLE") or "").lower()
                org_id = row.get("org_id") or ""
                group_id = row.get("group_id") or ""
                version = row.get("mongodb_version") or ""
                process_type = row.get("process_type") or ""
                val_int = int(row.get("Value") or row.get("value") or "0")
                job = row.get("job") or ""

                if not hostname:
                    continue

                dc_info = resolve_dc_from_mongo_hostname(hostname)
                dc = await get_or_create_dc(db, dc_info)

                is_primary_text = replica_state == "PRIMARY"
                is_primary_int = val_int == 1
                is_mongos = role == "mongos" or process_type == "mongos"
                is_config = "config" in role or process_type == "config"

                asset_env = "PRODUCTION" if env in ["PRODUCTION", "PROD"] else "DR" if env == "DR" else "UAT"
                agree = (is_primary_text == is_primary_int) or is_mongos or is_config
                conf = 3 if agree else 2
                has_conflict = not agree and replica_state != "NONE" and val_int != 0

                is_primary = is_primary_int if has_conflict else is_primary_text
                rep_role = "MONGOS" if is_mongos else "CONFIG_SVR" if is_config else "PRIMARY" if is_primary else "SECONDARY"

                app_id = "MONGO_INFRA"
                app_name = "MongoDB DB Tier"
                if job:
                    parts = re.split(r'[-_]', job)
                    if parts and parts[0]:
                        app_id = parts[0].upper()
                        app_name = f"{app_id} Database Tier"

                asset = RuntimeAsset(
                    id=str(uuid.uuid4()),
                    name=hostname,
                    asset_type="MONGO_NODE",
                    tech_stack="mongodb",
                    environment=asset_env,
                    host=hostname,
                    platform="LINUX",
                    data_center_short=dc.short_name,
                    latest_confidence_level=conf,
                    latest_operational_state="ACTIVE" if (is_primary or is_mongos) else "STANDBY",
                    latest_replication_role=rep_role,
                    write_authority=is_primary,
                    is_deterministic=True,
                    data_source="mongodb",
                    metadata_json={
                        "rs_nm": rs_nm,
                        "cluster": cluster,
                        "org_id": org_id,
                        "group_id": group_id,
                        "mongodb_version": version,
                        "process_type": process_type,
                        "value_int": str(val_int),
                        "internal_conflict": f"text={replica_state} vs int={val_int}" if has_conflict else "",
                        "application_id": app_id,
                        "application_name": app_name
                    }
                )
                db.add(asset)
                assets_created += 1

        elif source == "oracle_oem":
            for row in reader:
                role_name = (row.get("role_name") or row.get("ROLE_NAME") or "").upper()
                target_name = row.get("target_name") or row.get("TARGET_NAME")
                env = (row.get("env") or row.get("ENV") or "uat").upper()

                if not target_name:
                    continue

                host = target_name
                db_name = target_name
                if "@" in target_name:
                    db_name, host = target_name.split("@", 1)
                elif "_" in target_name:
                    parts = target_name.split("_")
                    host = parts[-1]
                    db_name = "_".join(parts[:-1])

                dc_info = resolve_dc_from_oracle_hostname(host)
                dc = await get_or_create_dc(db, dc_info)

                is_standby = "STANDBY" in role_name
                asset_env = "PRODUCTION" if env in ["PRODUCTION", "PROD"] else "DR" if env == "DR" else "UAT"

                app_id = "ORACLE_INFRA"
                app_name = "Oracle Core databases"
                if target_name:
                    parts = target_name.split("_")
                    if parts and parts[0]:
                        app_id = parts[0].upper()
                        app_name = f"{app_id} Database Services"

                asset = RuntimeAsset(
                    id=str(uuid.uuid4()),
                    name=db_name,
                    asset_type="ORACLE_DB",
                    tech_stack="oracle",
                    environment=asset_env,
                    host=host,
                    platform="UNIX",
                    data_center_short=dc.short_name,
                    latest_confidence_level=3,
                    latest_operational_state="STANDBY" if is_standby else "ACTIVE",
                    latest_replication_role="PHYSICAL_STANDBY" if is_standby else "PRIMARY",
                    write_authority=not is_standby,
                    is_deterministic=True,
                    data_source="oracle_oem",
                    metadata_json={
                        "role_name": role_name,
                        "target_name": target_name,
                        "application_id": app_id,
                        "application_name": app_name
                    }
                )
                db.add(asset)
                assets_created += 1

        elif source == "mssql":
            for row in reader:
                hostname = row.get("hostname") or row.get("HOSTNAME")
                ag_name = row.get("ag_name") or row.get("AG_NAME") or "MSSQL-AG"
                replica_role = (row.get("replica_role") or row.get("REPLICA_ROLE") or "PRIMARY").upper()
                sync_state = (row.get("sync_state") or row.get("SYNC_STATE") or "SYNCHRONIZED").upper()
                db_name = row.get("db_name") or row.get("DB_NAME") or "PatientCareDB"
                env = (row.get("env") or row.get("ENV") or "UAT").upper()

                if not hostname:
                    continue

                dc_info = resolve_dc_from_mssql_hostname(hostname)
                dc = await get_or_create_dc(db, dc_info)

                is_primary = replica_role == "PRIMARY"
                is_sync = sync_state == "SYNCHRONIZED"
                conf = 4 if is_sync else 3
                asset_env = "PRODUCTION" if env in ["PRODUCTION", "PROD"] else "DR" if env == "DR" else "UAT"

                app_id = "MSSQL_INFRA"
                app_name = "MSSQL Core databases"
                if db_name:
                    app_id = db_name.replace("DB", "").upper()
                    app_name = f"{app_id} Database Tier"

                asset = RuntimeAsset(
                    id=str(uuid.uuid4()),
                    name=db_name,
                    asset_type="DATABASE_INSTANCE",
                    tech_stack="mssql",
                    environment=asset_env,
                    host=hostname,
                    platform="WINDOWS",
                    data_center_short=dc.short_name,
                    latest_confidence_level=conf,
                    latest_operational_state="STANDBY" if not is_primary else "ACTIVE",
                    latest_replication_role=replica_role,
                    write_authority=is_primary,
                    is_deterministic=True,
                    data_source="mssql",
                    metadata_json={
                        "ag_name": ag_name,
                        "sync_state": sync_state,
                        "replica_role": replica_role,
                        "application_id": app_id,
                        "application_name": app_name
                    }
                )
                db.add(asset)
                assets_created += 1

        elif source == "kafka":
            for row in reader:
                hostname = row.get("hostname") or row.get("HOSTNAME")
                broker_id = row.get("broker_id") or row.get("BROKER_ID") or "1"
                is_controller_str = str(row.get("is_controller") or row.get("IS_CONTROLLER") or "false").lower()
                urp_str = str(row.get("under_replicated_partitions") or row.get("URP") or "0")
                env = (row.get("env") or row.get("ENV") or "UAT").upper()

                if not hostname:
                    continue

                dc_info = resolve_dc_from_kafka_hostname(hostname)
                dc = await get_or_create_dc(db, dc_info)

                is_controller = is_controller_str in ["true", "1", "yes"]
                urp = int(urp_str) if urp_str.isdigit() else 0
                conf = 4 if urp == 0 else 2
                asset_env = "PRODUCTION" if env in ["PRODUCTION", "PROD"] else "DR" if env == "DR" else "UAT"

                app_id = "KAFKA_INFRA"
                app_name = "Kafka Shared Cluster"

                asset = RuntimeAsset(
                    id=str(uuid.uuid4()),
                    name=f"kafka-broker-{broker_id}",
                    asset_type="MESSAGING_NODE",
                    tech_stack="kafka",
                    environment=asset_env,
                    host=hostname,
                    platform="LINUX",
                    data_center_short=dc.short_name,
                    latest_confidence_level=conf,
                    latest_operational_state="DEGRADED" if urp > 0 else "ACTIVE",
                    latest_replication_role="CONTROLLER" if is_controller else "BROKER",
                    write_authority=True,
                    is_deterministic=True,
                    data_source="kafka",
                    metadata_json={
                        "broker_id": broker_id,
                        "is_controller": is_controller,
                        "under_replicated_partitions": urp,
                        "application_id": app_id,
                        "application_name": app_name
                    }
                )
                db.add(asset)
                assets_created += 1

        elif source == "avi_loadbalancer":
            for row in reader:
                name = row.get("name") or row.get("hostname") or row.get("VIP_NAME") or "avi-vip"
                app_id = row.get("app_id") or row.get("APP_ID") or "AVI_INFRA"
                
                # Check for site (GSLB) or active_dc (standard)
                site = row.get("site") or row.get("active_dc") or row.get("ACTIVE_DC") or "IBB1"
                
                # Check for controller (GSLB) or hostname (standard)
                host = row.get("controller") or row.get("hostname") or row.get("VIP_NAME") or name
                
                pool = row.get("pool") or row.get("active_pool") or row.get("ACTIVE_POOL") or "default-pool"
                health_score_str = str(row.get("health_score") or row.get("HEALTH_SCORE") or "100")
                tenant = row.get("tenant") or ""
                
                env = "PRODUCTION"
                if "uat" in tenant.lower() or "dev" in tenant.lower():
                    env = "UAT"
                elif "env" in row:
                    env = row.get("env").upper()

                if not name:
                    continue

                # Resolve DC
                dc_info = {"name": f"DC {site.upper()}", "short_name": site.upper()}
                if site.upper() in ["IBB1", "SHV", "GA-PRD", "MA-PRD"]:
                    dc_info = resolve_dc_from_avi_hostname(host)
                dc = await get_or_create_dc(db, dc_info)

                health_score = int(health_score_str) if health_score_str.isdigit() else 100
                conf = 4 if health_score >= 90 else 3

                asset = RuntimeAsset(
                    id=str(uuid.uuid4()),
                    name=name,
                    asset_type="LOAD_BALANCER",
                    tech_stack="avi",
                    environment=env,
                    host=host,
                    platform="LINUX",
                    data_center_short=dc.short_name,
                    latest_confidence_level=conf,
                    latest_operational_state="ACTIVE",
                    latest_replication_role="ACTIVE",
                    write_authority=True,
                    is_deterministic=True,
                    data_source="avi_loadbalancer",
                    metadata_json={
                        "vip_ip": row.get("vip_ip") or "10.0.0.1",
                        "active_pool": pool,
                        "active_dc": dc.short_name,
                        "health_score": health_score,
                        "application_id": app_id,
                        "application_name": f"{app_id} Services",
                        "tenant": tenant,
                        "zone": row.get("zone") or "",
                        "neighborhood": row.get("neighborhood") or ""
                    }
                )
                db.add(asset)
                assets_created += 1

        elif source == "scom":
            for row in reader:
                replica_name = row.get("ReplicaName") or row.get("replica_name") or ""
                role = (row.get("Role") or row.get("role") or "Secondary").strip()
                health_state = (row.get("HealthState") or row.get("health_state") or "Success").strip()

                if not replica_name:
                    continue

                host = replica_name.split("\\")[0] if "\\" in replica_name else replica_name

                dc_info = resolve_dc_from_mssql_hostname(host)
                dc = await get_or_create_dc(db, dc_info)

                is_primary = role.lower() in ["primary", "standalone"]
                is_healthy = health_state.lower() in ["success", "healthy", "ok"]
                conf = 4 if (is_primary and is_healthy) else 3 if is_healthy else 2

                app_id = "SCOM_INFRA"
                app_name = "SCOM SQL Services"
                if replica_name and "\\" in replica_name:
                    db_part = replica_name.split("\\")[1]
                    app_id = db_part.split("_")[0].upper()
                    app_name = f"{app_id} SQL Instance"

                asset = RuntimeAsset(
                    id=str(uuid.uuid4()),
                    name=replica_name,
                    asset_type="DATABASE_INSTANCE",
                    tech_stack="mssql",
                    environment="PRODUCTION",
                    host=host,
                    platform="WINDOWS",
                    data_center_short=dc.short_name,
                    latest_confidence_level=conf,
                    latest_operational_state="ACTIVE" if is_primary else "STANDBY",
                    latest_replication_role="PRIMARY" if is_primary else "SECONDARY",
                    write_authority=is_primary,
                    is_deterministic=True,
                    data_source="scom",
                    metadata_json={
                        "health_state": health_state,
                        "role": role,
                        "application_id": app_id,
                        "application_name": app_name
                    }
                )
                db.add(asset)
                assets_created += 1

        elif source == "ocp":
            for row in reader:
                pod = row.get("pod") or row.get("POD") or ""
                namespace = row.get("namespace") or row.get("NAMESPACE") or ""
                cluster = row.get("cluster") or row.get("CLUSTER") or ""
                env = (row.get("env") or row.get("ENV") or "UAT").upper()
                lob = row.get("lob") or row.get("LOB") or ""
                nh = row.get("neighborhood") or row.get("NEIGHBORHOOD") or ""

                if not pod:
                    continue

                # Determine DC from cluster name prefix (e.g. dcgl... -> GL, dcms... -> MS)
                dc_prefix = "UNK"
                if cluster.startswith("dc"):
                    dc_prefix = cluster[2:4].upper()
                elif cluster:
                    dc_prefix = cluster[:3].upper()

                dc = await get_or_create_dc(db, {"name": f"OCP Cluster {dc_prefix}", "short_name": dc_prefix})
                asset_env = "PRODUCTION" if env in ["PRODUCTION", "PROD"] else "DR" if env == "DR" else "UAT"

                app_id = "OCP_INFRA"
                app_name = "OpenShift Cluster Services"
                if namespace:
                    app_id = namespace.upper()
                    app_name = f"{namespace.title()} Microservices"

                asset = RuntimeAsset(
                    id=str(uuid.uuid4()),
                    name=pod,
                    asset_type="OCP_POD",
                    tech_stack="ocp",
                    environment=asset_env,
                    host=cluster,
                    platform="LINUX",
                    data_center_short=dc.short_name,
                    latest_confidence_level=4,
                    latest_operational_state="ACTIVE",
                    latest_replication_role="NONE",
                    write_authority=False,
                    is_deterministic=True,
                    data_source="ocp",
                    metadata_json={
                        "namespace": namespace,
                        "lob": lob,
                        "neighborhood": nh,
                        "application_id": app_id,
                        "application_name": app_name
                    }
                )
                db.add(asset)
                assets_created += 1

        elif source == "appdynamics":
            for row in reader:
                machine_name = row.get("machine_name") or row.get("MACHINE_NAME") or ""
                app_full_name = row.get("app_full_name") or row.get("APP_FULL_NAME") or ""
                app_id_val = row.get("app_id") or row.get("APP_ID") or ""
                node_name = row.get("node_name") or row.get("NODE_NAME") or ""
                tier_name = row.get("tier_name") or row.get("TIER_NAME") or ""
                
                metric_path = row.get("metric_path") or ""
                if metric_path and not node_name:
                    if "Individual Nodes|" in metric_path:
                        parts = metric_path.split("Individual Nodes|")
                        if len(parts) > 1:
                            subparts = parts[1].split("|")
                            node_name = subparts[0]
                            machine_name = node_name
                    elif "Component:" in metric_path:
                        parts = metric_path.split("Component:")
                        if len(parts) > 1:
                            subparts = parts[1].split("|")
                            node_name = f"comp-{subparts[0]}"
                            machine_name = node_name

                if not node_name:
                    node_name = row.get("id") or "appd-node"
                    machine_name = node_name

                # Infer DC from prefix
                dc_short = "UNK"
                if machine_name:
                    m = machine_name.upper()
                    if m.startswith("STR"): dc_short = "STR"
                    elif m.startswith("GARD") or m.startswith("GAR"): dc_short = "GAR"
                    elif m.startswith("MAN"): dc_short = "MAN"
                    elif m.startswith("LEW"): dc_short = "LEW"
                    elif m.startswith("ARV"): dc_short = "ARV"
                    elif m.startswith("SHV"): dc_short = "SHV"
                    elif m.startswith("TPE"): dc_short = "TPE"
                    elif m.startswith("OXM"): dc_short = "OXM"
                    elif "PROD-1" in m:
                        parts = m.split("PROD")
                        if parts[0]:
                            dc_short = parts[0].replace("-", "").strip()[:4]
                
                dc = await get_or_create_dc(db, {"name": f"DC {dc_short}", "short_name": dc_short})
                
                asset = RuntimeAsset(
                    id=str(uuid.uuid4()),
                    name=node_name,
                    asset_type="COMPUTE_NODE",
                    tech_stack="appdynamics",
                    environment="PRODUCTION",
                    host=machine_name or node_name,
                    platform="LINUX",
                    data_center_short=dc.short_name,
                    latest_confidence_level=4,
                    latest_operational_state="ACTIVE",
                    latest_replication_role="NONE",
                    write_authority=True,
                    is_deterministic=True,
                    data_source="appdynamics",
                    metadata_json={
                        "application_id": app_id_val,
                        "application_name": app_full_name or f"{app_id_val} App",
                        "node_name": node_name,
                        "tier_name": tier_name,
                        "metric_name": row.get("metric_name") or "",
                        "metric_value": row.get("value") or ""
                    }
                )
                db.add(asset)
                assets_created += 1

        elif source == "batch":
            for row in reader:
                mach_name = row.get("MACH_NAME") or row.get("mach_name") or row.get("RUN_MACHINE") or ""
                job_name = row.get("JOB_NAME") or row.get("job_name") or ""
                instance = row.get("Instance") or row.get("INSTANCE") or ""
                as_application = row.get("AS_APPLICATION") or row.get("as_application") or row.get("AS_APPLIC") or row.get("as_applic") or ""
                job_status = (row.get("JOB_STATUS") or row.get("STATUS") or "").upper()

                if not job_name:
                    continue

                # Determine DC from machine name if possible
                dc_short = "UNK"
                if mach_name:
                    m = mach_name.upper()
                    if "EPVRA" in m or "EPV" in m:
                        dc_short = "EPV"
                    elif "EDA" in m:
                        dc_short = "EDA"
                    elif "MRB" in m:
                        dc_short = "MRB"

                dc = await get_or_create_dc(db, {"name": f"DC {dc_short}", "short_name": dc_short})

                app_id = as_application or "BATCH_INFRA"
                app_name = f"{app_id} Batch Workloads"

                asset = RuntimeAsset(
                    id=str(uuid.uuid4()),
                    name=job_name,
                    asset_type="BATCH_JOB",
                    tech_stack="batch",
                    environment="PRODUCTION",
                    host=mach_name,
                    platform="LINUX",
                    data_center_short=dc.short_name,
                    latest_confidence_level=4,
                    latest_operational_state="ACTIVE" if job_status == "SUCCESS" else "STANDBY",
                    latest_replication_role="NONE",
                    write_authority=False,
                    is_deterministic=True,
                    data_source="batch",
                    metadata_json={
                        "job_name": job_name,
                        "instance": instance,
                        "as_application": as_application,
                        "job_status": job_status,
                        "application_id": app_id,
                        "application_name": app_name
                    }
                )
                db.add(asset)
                assets_created += 1

        else:  # CMDB / SPLOC / fallback
            # Check if this is a SPLOC traffic file
            is_sploc = "wf_dc" in reader.fieldnames and "app_id" in reader.fieldnames
            if is_sploc:
                for row in reader:
                    app_id = row.get("app_id") or "UNKNOWN"
                    app_name = f"{app_id} Application"
                    dc_short = row.get("wf_dc") or "UNK"
                    service = row.get("sf_service") or row.get("wf_acln") or "unknown-service"
                    avg_value = row.get("avg_value") or "0"
                    total_value = row.get("total_value") or "0"
                    sample_count = row.get("sample_count") or "0"

                    if not service or not app_id:
                        continue

                    dc = await get_or_create_dc(db, {"name": f"DC {dc_short}", "short_name": dc_short})

                    asset = RuntimeAsset(
                        id=str(uuid.uuid4()),
                        name=service,
                        asset_type="COMPUTE_NODE",
                        tech_stack="java",
                        environment="PRODUCTION",
                        host=f"{service.lower()}.{dc_short.lower()}.healthmesh.ai",
                        platform="LINUX",
                        data_center_short=dc.short_name,
                        latest_confidence_level=4,
                        latest_operational_state="ACTIVE",
                        latest_replication_role="NONE",
                        write_authority=True,
                        is_deterministic=True,
                        data_source="sploc",
                        metadata_json={
                            "application_id": app_id,
                            "application_name": app_name,
                            "avg_latency_ms": avg_value,
                            "request_count": total_value,
                            "sample_count": sample_count,
                            "service_name": service
                        }
                    )
                    db.add(asset)
                    assets_created += 1
            else:
                for row in reader:
                    app_name = row.get("APPLICATION_NAME") or row.get("application_name") or ""
                    app_id = row.get("APPLICATION_ID") or row.get("application_id") or app_name.split(" ")[0].upper()
                    env = (row.get("ENVIRONMENT") or row.get("environment") or "UAT").upper()
                    device_name = row.get("DEVICE_NAME") or row.get("device_name")
                    device_type = row.get("DEVICE_TYPE") or row.get("device_type") or "SERVER"
                    dc_name = row.get("DATA_CENTER") or row.get("data_center") or "Unknown DC"

                    lvl1_name = row.get("DEVICE_LVL1_NAME") or ""
                    lvl1_type = row.get("DEVICE_LVL1_TYPE") or ""
                    lvl2_name = row.get("DEVICE_LVL2_NAME") or ""
                    lvl2_type = row.get("DEVICE_LVL2_TYPE") or ""
                    lvl3_name = row.get("DEVICE_LVL3_NAME") or ""
                    lvl3_type = row.get("DEVICE_LVL3_TYPE") or ""
                    lvl4_name = row.get("DEVICE_LVL4_NAME") or ""
                    lvl4_type = row.get("DEVICE_LVL4_TYPE") or ""

                    if not device_name or not app_name:
                        continue

                    # Classify tech stack
                    t = device_type.upper()
                    stack = "vm"
                    if "MQ" in t or "QUEUE" in t:
                        stack = "ibm_mq"
                    elif "MONGO" in t:
                        stack = "mongodb"
                    elif "ORACLE" in t or "ORA" in t:
                        stack = "oracle"
                    elif "SQL" in t:
                        stack = "mssql"
                    elif "KAFKA" in t:
                        stack = "kafka"
                    elif "OCP" in t or "POD" in t or "KUBE" in t:
                        stack = "ocp"

                    # Classify asset type
                    asset_type = "SERVER"
                    if "MQ" in t:
                        asset_type = "MQ_QMGR"
                    elif "MONGO" in t:
                        asset_type = "MONGO_NODE"
                    elif "ORACLE" in t:
                        asset_type = "ORACLE_DB"
                    elif "OCP" in t or "POD" in t:
                        asset_type = "OCP_POD"

                    dc_short = dc_name.replace("DC ", "").replace(" ", "-").upper()[:8] or "UNK"
                    dc = await get_or_create_dc(db, {"name": dc_name, "short_name": dc_short})

                    asset_env = "PRODUCTION" if env in ["PRODUCTION", "PROD"] else "DR" if env == "DR" else "UAT"

                    asset = RuntimeAsset(
                        id=str(uuid.uuid4()),
                        name=device_name,
                        asset_type=asset_type,
                        tech_stack=stack,
                        environment=asset_env,
                        host=device_name,
                        platform="LINUX",
                        data_center_short=dc.short_name,
                        latest_confidence_level=4,
                        latest_operational_state="ACTIVE",
                        latest_replication_role="NONE",
                        write_authority=False,
                        is_deterministic=True,
                        data_source="cmdb",
                        metadata_json={
                            "application_id": app_id,
                            "application_name": app_name,
                            "device_type": device_type,
                            "lvl1": f"{lvl1_name}({lvl1_type})" if lvl1_name else "",
                            "lvl2": f"{lvl2_name}({lvl2_type})" if lvl2_name else "",
                            "lvl3": f"{lvl3_name}({lvl3_type})" if lvl3_name else "",
                            "lvl4": f"{lvl4_name}({lvl4_type})" if lvl4_name else "",
                            "device_chain": " → ".join(filter(None, [lvl1_name, lvl2_name, lvl3_name, lvl4_name]))
                        }
                    )
                    db.add(asset)
                    assets_created += 1

    except Exception as e:
        logger.error(f"Error parsing CSV content: {e}")
        errors.append(str(e))

    status = "FAILED" if errors else "SUCCESS"
    return {
        "assets_created": assets_created,
        "errors": errors,
        "source": source,
        "status": status
    }

@router.post("/import", response_model=Dict[str, Any])
async def import_csv(
    file: UploadFile = File(...),
    source_type: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db)
):
    import io
    import csv
    import uuid
    from datetime import datetime
    
    content_bytes = await file.read()
    
    if file.filename.endswith(".xlsx"):
        try:
            import openpyxl
            f_buf = io.BytesIO(content_bytes)
            wb = openpyxl.load_workbook(f_buf, read_only=True, data_only=True)
            ws = wb.active
            rows_list = list(ws.iter_rows(values_only=True))
            wb.close()
            if len(rows_list) < 2:
                return {
                    "id": str(uuid.uuid4()),
                    "source_name": source_type or "xlsx",
                    "file_name": file.filename,
                    "imported_at": datetime.utcnow().isoformat() + "Z",
                    "record_count": 0,
                    "status": "FAILED",
                    "errors": ["XLSX file is empty"]
                }
            headers = [str(h or '').strip() for h in rows_list[0]]
            csv_buf = io.StringIO()
            writer = csv.writer(csv_buf)
            writer.writerow(headers)
            for data_row in rows_list[1:]:
                writer.writerow([str(c or '') for c in data_row])
            content = csv_buf.getvalue()
            csv_fname = file.filename.replace(".xlsx", ".csv")
            result = await parse_and_insert_csv(csv_fname, content, source_type, db)
        except ImportError:
            result = {
                "assets_created": 0,
                "errors": ["openpyxl not installed on backend"],
                "source": source_type or "xlsx",
                "status": "FAILED"
            }
        except Exception as e:
            result = {
                "assets_created": 0,
                "errors": [str(e)],
                "source": source_type or "xlsx",
                "status": "FAILED"
            }
    elif file.filename.endswith(".json"):
        content = content_bytes.decode("utf-8", errors="ignore")
        result = await parse_and_insert_json_file(file.filename, content, db)
    else:
        content = content_bytes.decode("utf-8", errors="ignore")
        result = await parse_and_insert_csv(file.filename, content, source_type, db)
        
    assets_created = result["assets_created"]
    errors = result["errors"]
    source = result["source"]
    status = result["status"]
    
    # Save import history log
    imp = DataSourceImport(
        id=str(uuid.uuid4()),
        source_name=source,
        file_name=file.filename,
        record_count=assets_created,
        status=status,
        errors=errors
    )
    db.add(imp)
    
    # Save Audit log entry
    audit = RuntimeAuditLog(
        id=str(uuid.uuid4()),
        event_type="IMPORT",
        description=f"Imported {assets_created} records from {file.filename} ({source})" if not errors else f"Failed import of {file.filename}: {errors[0]}",
        source=source
    )
    db.add(audit)
    
    await db.commit()
    
    return {
        "id": imp.id,
        "source_name": source,
        "file_name": file.filename,
        "imported_at": imp.imported_at.isoformat() + "Z",
        "record_count": assets_created,
        "status": status,
        "errors": errors
    }

async def parse_and_insert_json_file(file_name: str, content: str, db: AsyncSession) -> Dict[str, Any]:
    import json
    import re
    import uuid
    
    assets_created = 0
    errors = []
    source = "json_unsupported"
    
    try:
        is_json = False
        try:
            data = json.loads(content)
            is_json = True
        except json.JSONDecodeError:
            pass
            
        if is_json:
            if isinstance(data, dict) and (data.get("style") == "flexvol" or "svm" in data):
                source = "netapp"
                name = data.get("name") or "datavol"
                svm_name = data.get("svm", {}).get("name") or "vsdata"
                state = data.get("state") or "online"
                vol_type = data.get("type") or "rw"
                uuid_str = data.get("uuid") or str(uuid.uuid4())
                
                dc = await get_or_create_dc(db, {"name": "DC Birmingham IBB1", "short_name": "IBB1"})
                
                asset = RuntimeAsset(
                    id=str(uuid.uuid4()),
                    name=name,
                    asset_type="STORAGE_VOLUME",
                    tech_stack="netapp",
                    environment="PRODUCTION",
                    host=svm_name,
                    platform="ONTAP",
                    data_center_short=dc.short_name,
                    latest_confidence_level=4,
                    latest_operational_state="ACTIVE" if state.lower() == "online" else "INACTIVE",
                    latest_replication_role="NONE",
                    write_authority=True if vol_type.lower() == "rw" else False,
                    is_deterministic=True,
                    data_source="netapp",
                    metadata_json={
                        "uuid": uuid_str,
                        "comment": data.get("comment") or "",
                        "size_bytes": data.get("size") or 0,
                        "style": data.get("style") or "",
                        "type": vol_type,
                        "application_id": "NETAPP",
                        "application_name": "NetApp Storage Infrastructure"
                    }
                )
                db.add(asset)
                assets_created += 1
            else:
                source = "generic_json"
        else:
            # Parse Prometheus exposition metrics
            prometheus_lines = content.strip().split("\n")
            resource_metrics = {}
            for line in prometheus_lines:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                
                match = re.match(r"^([a-zA-Z0-9_]+)\{(.*)\}\s+([0-9e\.\+\-]+)", line)
                if not match:
                    continue
                
                metric_name, label_str, val_str = match.groups()
                val = float(val_str)
                
                labels = {}
                for lbl_match in re.finditer(r'([a-zA-Z0-9_]+)="([^"]*)"', label_str):
                    k, v = lbl_match.groups()
                    labels[k] = v
                
                uuid_lbl = labels.get("uuid")
                name_lbl = labels.get("name")
                if not uuid_lbl or not name_lbl:
                    continue
                
                server_lbl = labels.get("server") or ""
                group_key = (uuid_lbl, server_lbl)
                
                if group_key not in resource_metrics:
                    resource_metrics[group_key] = {
                        "labels": labels,
                        "metrics": {}
                    }
                resource_metrics[group_key]["metrics"][metric_name] = val
                
            if resource_metrics:
                source = "avi_loadbalancer"
                for (uuid_lbl, server_lbl), info in resource_metrics.items():
                    labels = info["labels"]
                    metrics = info["metrics"]
                    
                    name_lbl = labels.get("name")
                    type_lbl = labels.get("type") or "pool"
                    tenant = labels.get("tenant") or "adcs"
                    
                    health_score = metrics.get("avi_healthscore_health_score_value", 100.0)
                    performance_score = metrics.get("avi_healthscore_performance_score", 100.0)
                    
                    if server_lbl:
                        asset_name = f"{name_lbl}-{server_lbl}"
                        host_part = server_lbl.split(":")[0] if ":" in server_lbl else server_lbl
                        port_part = int(server_lbl.split(":")[1]) if ":" in server_lbl else None
                    else:
                        asset_name = name_lbl
                        host_part = f"{name_lbl}.healthmesh.ai"
                        port_part = 443
                        
                    dc_info = resolve_dc_from_avi_hostname(host_part)
                    if host_part == "10.1.1.1":
                        dc_info = {"name": "DC Birmingham IBB1", "short_name": "IBB1"}
                    elif host_part == "30.0.60.109":
                        dc_info = {"name": "DC Shoreview", "short_name": "SHV"}
                        
                    dc = await get_or_create_dc(db, dc_info)
                    
                    env = "PRODUCTION"
                    if "uat" in tenant.lower() or "dev" in tenant.lower():
                        env = "UAT"
                        
                    op_state = "ACTIVE"
                    if health_score == 0.0:
                        op_state = "DEGRADED"
                        
                    asset = RuntimeAsset(
                        id=str(uuid.uuid4()),
                        name=asset_name,
                        asset_type="LOAD_BALANCER",
                        tech_stack="avi",
                        environment=env,
                        host=host_part,
                        port=port_part,
                        platform="LINUX",
                        data_center_short=dc.short_name,
                        latest_confidence_level=4,
                        latest_operational_state=op_state,
                        latest_replication_role="ACTIVE" if type_lbl == "virtualservice" else "NONE",
                        write_authority=True if type_lbl == "virtualservice" else False,
                        is_deterministic=True,
                        data_source="avi_loadbalancer",
                        metadata_json={
                            "uuid": uuid_lbl,
                            "type": type_lbl,
                            "tenant": tenant,
                            "name": name_lbl,
                            "server": server_lbl,
                            "health_score": health_score,
                            "performance_score": performance_score,
                            "application_id": "ADCS",
                            "application_name": "ADCS Load Balancing Services"
                        }
                    )
                    db.add(asset)
                    assets_created += 1
                    
    except Exception as e:
        logger.error(f"Error parsing JSON/Metric file: {e}")
        errors.append(str(e))
        
    return {
        "assets_created": assets_created,
        "errors": errors,
        "source": source,
        "status": "FAILED" if errors else "SUCCESS"
    }

@router.post("/import-all-docs", response_model=Dict[str, Any])
async def import_all_docs(db: AsyncSession = Depends(get_db)):
    import os
    from fastapi import HTTPException
    
    # Locate docs directory first
    docs_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../../docs"))
    if not os.path.exists(docs_dir):
        docs_dir = os.path.abspath(os.path.join(os.getcwd(), "backend", "docs"))
        if not os.path.exists(docs_dir):
            docs_dir = os.path.abspath(os.path.join(os.getcwd(), "docs"))
            
    if not os.path.exists(docs_dir):
        raise HTTPException(status_code=404, detail=f"Documentation directory not found. Checked: {docs_dir}")
        
    # Clean all data inside the transaction (do not commit early, rollback on failure)
    await db.execute(delete(RuntimeAsset))
    await db.execute(delete(RuntimeDataCenter))
    await db.execute(delete(DataSourceImport))
    await db.execute(delete(RuntimeAuditLog))
    await db.execute(delete(SourceProposal))
    await db.execute(delete(ApplicationIntent))
        
    imported_files = []
    total_assets = 0
    errors = []
    
    # Get all CSV, JSON, and XLSX files in docs
    files = [f for f in os.listdir(docs_dir) if f.endswith(".csv") or f.endswith(".json") or f.endswith(".xlsx")]
    
    # Sort files
    for fname in sorted(files):
        fpath = os.path.join(docs_dir, fname)
        try:
            if fname.endswith(".xlsx"):
                # Convert XLSX to CSV in-memory using openpyxl
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                    ws = wb.active
                    rows_list = list(ws.iter_rows(values_only=True))
                    wb.close()
                    if len(rows_list) < 2:
                        logger.info(f"Skipping empty XLSX file: {fname}")
                        continue
                    headers = [str(h or '').strip() for h in rows_list[0]]
                    csv_buf = io.StringIO()
                    writer = csv.writer(csv_buf)
                    writer.writerow(headers)
                    for data_row in rows_list[1:]:
                        writer.writerow([str(c or '') for c in data_row])
                    content = csv_buf.getvalue()
                    csv_fname = fname.replace(".xlsx", ".csv")
                    result = await parse_and_insert_csv(csv_fname, content, None, db)
                except ImportError:
                    logger.warning(f"openpyxl not installed — skipping XLSX file: {fname}")
                    errors.append(f"{fname}: openpyxl not installed")
                    continue
                except Exception as xlsx_exc:
                    logger.error(f"Failed to parse XLSX file {fname}: {xlsx_exc}")
                    errors.append(f"{fname}: {str(xlsx_exc)}")
                    continue
            else:
                with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read()
                
                if fname.endswith(".json"):
                    result = await parse_and_insert_json_file(fname, content, db)
                else:
                    result = await parse_and_insert_csv(fname, content, None, db)
                
            assets_created = result["assets_created"]
            status = result["status"]
            file_errors = result["errors"]
            source = result["source"]
            
            if file_errors:
                errors.extend(file_errors)
                
            # Log individual imports
            imp = DataSourceImport(
                id=str(uuid.uuid4()),
                source_name=source,
                file_name=fname,
                record_count=assets_created,
                status=status,
                errors=file_errors
            )
            db.add(imp)
            total_assets += assets_created
            imported_files.append({
                "file": fname,
                "source": source,
                "count": assets_created,
                "status": status
            })
        except Exception as file_exc:
            logger.error(f"Failed to import file {fname}: {file_exc}")
            errors.append(f"{fname}: {str(file_exc)}")
            
    # 2. Automatically generate design intents for all discovered applications!
    result = await db.execute(select(RuntimeAsset))
    all_assets = result.scalars().all()
    
    apps_metadata = {}
    for asset in all_assets:
        app_id = "INFRASTRUCTURE"
        app_name = "Infrastructure Services"
        if asset.metadata_json and asset.metadata_json.get("application_id"):
            app_id = asset.metadata_json["application_id"]
            app_name = asset.metadata_json.get("application_name", app_id)
            
        if app_id not in apps_metadata:
            apps_metadata[app_id] = {
                "id": app_id,
                "name": app_name,
                "dcs": set(),
                "stacks": set()
            }
        
        if asset.data_center_short:
            apps_metadata[app_id]["dcs"].add(asset.data_center_short)
        if asset.tech_stack:
            apps_metadata[app_id]["stacks"].add(asset.tech_stack)
            
    # Create an ApplicationIntent for each discovered application
    for app_id, meta in apps_metadata.items():
        if app_id in ["INFRASTRUCTURE", "MQ_INFRA", "MONGO_INFRA", "ORACLE_INFRA", "SCOM_INFRA", "OCP_INFRA", "BATCH_INFRA"]:
            continue
            
        dcs_list = list(meta["dcs"])
        primary_dc = dcs_list[0] if dcs_list else "UNK"
        
        write_assets = [a for a in all_assets if a.metadata_json and a.metadata_json.get("application_id") == app_id and a.write_authority]
        if write_assets and write_assets[0].data_center_short:
            primary_dc = write_assets[0].data_center_short
            
        intent = ApplicationIntent(
            application_id=app_id,
            application_name=meta["name"],
            intended_active_dcs=dcs_list,
            intended_primary_dc=primary_dc,
            intended_environments=["PRODUCTION"],
            failover_type="AUTOMATIC",
            replication_model="SINGLE_WRITER" if len(dcs_list) > 1 else "STANDALONE",
            required_tech_stacks=list(meta["stacks"]),
            alignment_status="UNKNOWN"
        )
        db.add(intent)
        
    await db.commit()
    
    # 3. Compute alignment status for all applications by running drift detection
    result_intents = await db.execute(select(ApplicationIntent))
    intents = result_intents.scalars().all()
    for intent in intents:
        drifts = await run_drift_detection(db, intent.application_id, environment="PRODUCTION", persist_critical=True)
        intent.alignment_status = compute_alignment_status(drifts)
        db.add(intent)
        
    # Audit log
    audit = RuntimeAuditLog(
        id=str(uuid.uuid4()),
        event_type="BULK_IMPORT",
        description=f"Operator executed one-shot bulk import from docs folder. Processed {len(imported_files)} files. Created {total_assets} assets, {len(intents)} application design intents."
    )
    db.add(audit)
    await db.commit()
    
    return {
        "status": "SUCCESS" if not errors else "PARTIAL",
        "message": f"Successfully processed {len(imported_files)} documentation files.",
        "imported_files": imported_files,
        "total_assets": total_assets,
        "total_intents": len(intents),
        "errors": errors
    }

@router.post("/seed", response_model=Dict[str, Any])
async def seed_data(db: AsyncSession = Depends(get_db)):
    return await import_all_docs(db)

@router.post("/reset", response_model=Dict[str, Any])
async def reset_data(db: AsyncSession = Depends(get_db)):
    await db.execute(delete(RuntimeAsset))
    await db.execute(delete(RuntimeDataCenter))
    await db.execute(delete(DataSourceImport))
    await db.execute(delete(RuntimeAuditLog))
    await db.execute(delete(SourceProposal))
    await db.execute(delete(ApplicationIntent))
    
    # Audit log
    aud = RuntimeAuditLog(
        id=str(uuid.uuid4()),
        event_type="SYSTEM_RESET",
        description="Operator initiated a system reset: all database records cleared."
    )
    db.add(aud)
    await db.commit()
    
    return {"status": "SUCCESS", "message": "All runtime tables reset successfully."}

@router.get("/audit-logs", response_model=List[Dict[str, Any]])
async def get_audit_logs(application_id: Optional[str] = Query(None), db: AsyncSession = Depends(get_db)):
    query = select(RuntimeAuditLog)
    if application_id:
        query = query.where(RuntimeAuditLog.application_id == application_id)
    
    query = query.order_by(RuntimeAuditLog.occurred_at.desc())
    result = await db.execute(query)
    logs = result.scalars().all()

    return [
        {
            "id": log.id,
            "event_type": log.event_type,
            "description": log.description,
            "actor": log.actor,
            "source": log.source,
            "application_id": log.application_id,
            "occurred_at": log.occurred_at.isoformat() + "Z"
        } for log in logs
    ]

# ─── Intents management ──────────────────────────────────────────────────────

@router.get("/intents", response_model=List[Dict[str, Any]])
async def get_intents(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(ApplicationIntent))
    intents = result.scalars().all()
    return [
        {
            "application_id": i.application_id,
            "application_name": i.application_name,
            "intended_active_dcs": i.intended_active_dcs,
            "intended_primary_dc": i.intended_primary_dc,
            "intended_environments": i.intended_environments,
            "failover_type": i.failover_type,
            "replication_model": i.replication_model,
            "required_tech_stacks": i.required_tech_stacks,
            "created_at": i.created_at.isoformat() + "Z",
            "updated_at": i.updated_at.isoformat() + "Z"
        } for i in intents
    ]

@router.post("/intents", response_model=Dict[str, Any])
async def save_intent(intent_data: Dict[str, Any], db: AsyncSession = Depends(get_db)):
    app_id = intent_data["application_id"]
    result = await db.execute(select(ApplicationIntent).where(ApplicationIntent.application_id == app_id))
    existing = result.scalar_one_or_none()

    if existing:
        existing.intended_active_dcs = intent_data["intended_active_dcs"]
        existing.intended_primary_dc = intent_data["intended_primary_dc"]
        existing.intended_environments = intent_data.get("intended_environments", ["PRODUCTION"])
        existing.failover_type = intent_data.get("failover_type", "AUTOMATIC")
        existing.replication_model = intent_data["replication_model"]
        existing.required_tech_stacks = intent_data["required_tech_stacks"]
        existing.updated_at = datetime.utcnow()
        action = "INTENT_UPDATED"
        desc = f"Intent updated for {existing.application_name}"
    else:
        existing = ApplicationIntent(
            application_id=app_id,
            application_name=intent_data["application_name"],
            intended_active_dcs=intent_data["intended_active_dcs"],
            intended_primary_dc=intent_data["intended_primary_dc"],
            intended_environments=intent_data.get("intended_environments", ["PRODUCTION"]),
            failover_type=intent_data.get("failover_type", "AUTOMATIC"),
            replication_model=intent_data["replication_model"],
            required_tech_stacks=intent_data["required_tech_stacks"]
        )
        db.add(existing)
        action = "INTENT_CREATED"
        desc = f"Intent created for {existing.application_name}"

    audit = RuntimeAuditLog(
        id=str(uuid.uuid4()),
        event_type=action,
        description=desc,
        application_id=app_id
    )
    db.add(audit)
    
    await db.commit()
    
    return {"status": "SUCCESS", "message": "Design intent saved successfully."}

@router.delete("/intents/{app_id}", response_model=Dict[str, Any])
async def delete_intent(app_id: str, db: AsyncSession = Depends(get_db)):
    await db.execute(delete(ApplicationIntent).where(ApplicationIntent.application_id == app_id))
    
    audit = RuntimeAuditLog(
        id=str(uuid.uuid4()),
        event_type="INTENT_DELETED",
        description=f"Intent deleted for {app_id}",
        application_id=app_id
    )
    db.add(audit)
    
    await db.commit()
    
    return {"status": "SUCCESS", "message": f"Design intent deleted for {app_id}."}

# ─── Collaborative Proposals management ──────────────────────────────────────

@router.get("/proposals", response_model=List[Dict[str, Any]])
async def get_proposals(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(SourceProposal).order_by(SourceProposal.proposed_at.desc()))
    props = result.scalars().all()
    return [
        {
            "id": p.id,
            "source_name": p.source_name,
            "system": p.system,
            "signal_type": p.signal_type,
            "tech_stack": p.tech_stack,
            "rationale": p.rationale,
            "is_deterministic_claim": p.is_deterministic_claim,
            "proposed_by": p.proposed_by,
            "proposed_at": p.proposed_at.isoformat() + "Z",
            "status": p.status
        } for p in props
    ]

@router.post("/proposals", response_model=Dict[str, Any])
async def submit_proposal(p_data: Dict[str, Any], db: AsyncSession = Depends(get_db)):
    prop = SourceProposal(
        id=str(uuid.uuid4()),
        source_name=p_data["source_name"],
        system=p_data["system"],
        signal_type=p_data["signal_type"],
        tech_stack=p_data["tech_stack"],
        rationale=p_data["rationale"],
        is_deterministic_claim=p_data.get("is_deterministic_claim", True),
        proposed_by=p_data.get("proposed_by", "operator")
    )
    db.add(prop)

    audit = RuntimeAuditLog(
        id=str(uuid.uuid4()),
        event_type="PROPOSAL_SUBMITTED",
        description=f"Data source proposal submitted: {prop.source_name} ({prop.tech_stack})"
    )
    db.add(audit)
    
    await db.commit()
    
    return {"status": "SUCCESS", "id": prop.id}

@router.put("/proposals/{id}", response_model=Dict[str, Any])
async def update_proposal_status(id: str, status_data: Dict[str, Any], db: AsyncSession = Depends(get_db)):
    status = status_data["status"]
    await db.execute(update(SourceProposal).where(SourceProposal.id == id).values(status=status))
    
    audit = RuntimeAuditLog(
        id=str(uuid.uuid4()),
        event_type="PROPOSAL_UPDATED",
        description=f"Data source proposal status updated to {status} for ID {id}."
    )
    db.add(audit)
    
    await db.commit()
    
    return {"status": "SUCCESS"}

@router.get("/drift/{app_id}", response_model=List[Dict[str, Any]])
async def get_drift_for_app(
    app_id: str,
    environment: str = Query("PRODUCTION"),
    db: AsyncSession = Depends(get_db),
):
    """Run drift detection for a single application and return drift items."""
    drifts = await run_drift_detection(db, app_id, environment, persist_critical=True)
    # Update alignment_status on the intent
    alignment = compute_alignment_status(drifts)
    await db.execute(
        update(ApplicationIntent)
        .where(ApplicationIntent.application_id == app_id)
        .values(alignment_status=alignment)
    )
    await db.commit()
    return drifts


@router.get("/drift", response_model=Dict[str, Any])
async def get_drift_all(
    environment: str = Query("PRODUCTION"),
    db: AsyncSession = Depends(get_db),
):
    """Run drift detection for all applications that have an intent and return all drifts."""
    all_drifts = await run_drift_detection_all(db, environment)
    # Update alignment_status for each app
    for app_id, drifts in all_drifts.items():
        alignment = compute_alignment_status(drifts)
        await db.execute(
            update(ApplicationIntent)
            .where(ApplicationIntent.application_id == app_id)
            .values(alignment_status=alignment)
        )
    # For apps with no drifts, mark as ALIGNED if intent exists
    intent_res = await db.execute(select(ApplicationIntent))
    intents = intent_res.scalars().all()
    for intent in intents:
        if intent.application_id not in all_drifts:
            await db.execute(
                update(ApplicationIntent)
                .where(ApplicationIntent.application_id == intent.application_id)
                .values(alignment_status="ALIGNED")
            )
    await db.commit()
    total = sum(len(v) for v in all_drifts.values())
    return {"drifts": all_drifts, "total_drift_count": total}


@router.post("/conflicts/resolve", response_model=Dict[str, Any])
async def resolve_conflict(data: Dict[str, Any], db: AsyncSession = Depends(get_db)):
    asset_name = data.get("asset_name")
    authoritative_source = data.get("authoritative_source")
    
    if not asset_name or not authoritative_source:
        raise HTTPException(status_code=400, detail="Missing asset_name or authoritative_source")

    # Fetch all assets with this name or host matching this asset name
    result = await db.execute(
        select(RuntimeAsset).where(
            (RuntimeAsset.name == asset_name) | (RuntimeAsset.host == asset_name)
        )
    )
    group = result.scalars().all()
    
    if not group:
        raise HTTPException(status_code=404, detail="Asset not found")

    # Find the authoritative asset
    auth_asset = next((a for a in group if a.data_source == authoritative_source), None)
    if not auth_asset:
        # Fallback to the first asset if exact source match is not found
        auth_asset = group[0]

    # Align all other assets in the group to match the authoritative one
    for asset in group:
        if asset.id != auth_asset.id:
            asset.latest_replication_role = auth_asset.latest_replication_role
            asset.latest_operational_state = auth_asset.latest_operational_state
            asset.data_center_short = auth_asset.data_center_short
            asset.write_authority = auth_asset.write_authority
            asset.latest_confidence_level = 4  # Confirmed by operator override
            # Clear internal conflicts if any
            if asset.metadata_json and "internal_conflict" in asset.metadata_json:
                # Need to update metadata_json dict
                meta = dict(asset.metadata_json)
                meta["internal_conflict"] = ""
                asset.metadata_json = meta
    
    # Also log this in the audit log
    audit = RuntimeAuditLog(
        id=str(uuid.uuid4()),
        event_type="CONFLICT_RESOLVED",
        description=f"Operator manually resolved conflict for asset {asset_name}. Selected authoritative source: {authoritative_source}",
        source=authoritative_source
    )
    db.add(audit)
    
    await db.commit()

    return {"status": "SUCCESS", "message": f"Conflict for {asset_name} resolved successfully."}


# ─── Blast Radius / Failover Simulation ──────────────────────────────────────

@router.post("/simulate-failover", response_model=Dict[str, Any])
async def simulate_failover(
    data: Dict[str, Any],
    db: AsyncSession = Depends(get_db),
):
    """
    POST /api/v1/runtime-location/simulate-failover

    Calculate blast radius for a DC going offline.
    Body: { "dc": "IBB1" }
    Returns full impact analysis with critical/warning app lists and failover targets.
    """
    from app.services.blast_radius_service import calculate_blast_radius, BlastRadiusResult

    dc_name = data.get("dc")
    if not dc_name:
        raise HTTPException(status_code=400, detail="Missing 'dc' field in request body")

    result: BlastRadiusResult = await calculate_blast_radius(dc_name, db)

    return {
        "dc": result.dc,
        "dc_full_name": result.dc_full_name,
        "simulated_at": result.simulated_at,
        "total_apps_impacted": result.total_apps_impacted,
        "critical_count": result.critical_count,
        "warning_count": result.warning_count,
        "estimated_recovery_summary": result.estimated_recovery_summary,
        "failover_targets": result.failover_targets,
        "critical_apps": [
            {
                "application_id": a.application_id,
                "application_name": a.application_name,
                "environment": a.environment,
                "primary_dc": a.primary_dc,
                "has_failover": a.has_failover,
                "failover_target": a.failover_target,
                "promotion_required": a.promotion_required,
                "critical_reason": a.critical_reason,
                "affected_tech_stacks": a.affected_tech_stacks,
                "standby_dc": a.standby_dc,
            }
            for a in result.critical_apps
        ],
        "warning_apps": [
            {
                "application_id": a.application_id,
                "application_name": a.application_name,
                "environment": a.environment,
                "primary_dc": a.primary_dc,
                "has_failover": a.has_failover,
                "failover_target": a.failover_target,
                "promotion_required": a.promotion_required,
                "critical_reason": a.critical_reason,
                "affected_tech_stacks": a.affected_tech_stacks,
                "standby_dc": a.standby_dc,
            }
            for a in result.warning_apps
        ],
    }


@router.post("/failover", response_model=Dict[str, Any])
async def execute_failover(
    data: Dict[str, Any],
    db: AsyncSession = Depends(get_db),
):
    """
    POST /api/v1/runtime-location/failover
    Body: { "application_id": "PCA", "failed_dc": "IBB1", "promoted_dc": "SHV", "environment": "PRODUCTION" }
    """
    app_id = data.get("application_id")
    failed_dc = data.get("failed_dc")
    promoted_dc = data.get("promoted_dc")
    environment = data.get("environment", "PRODUCTION")

    if not app_id or not failed_dc or not promoted_dc:
        raise HTTPException(status_code=400, detail="Missing required fields: application_id, failed_dc, promoted_dc")

    # Fetch all assets for this app and environment
    result = await db.execute(select(RuntimeAsset))
    all_assets = result.scalars().all()

    app_assets = []
    for a in all_assets:
        if a.environment == environment:
            if a.metadata_json and a.metadata_json.get("application_id") == app_id:
                app_assets.append(a)
            elif app_id == "MQ_INFRA" and a.data_source == "ibm_mq":
                app_assets.append(a)
            elif app_id == "MONGO_INFRA" and a.data_source == "mongodb":
                app_assets.append(a)
            elif app_id == "ORACLE_INFRA" and a.data_source == "oracle_oem":
                app_assets.append(a)

    if not app_assets:
        raise HTTPException(status_code=404, detail="No assets found for the specified application and environment")

    # Mutate operational states and roles
    mutated_count = 0
    for asset in app_assets:
        # If the asset resides in the failed DC, make it offline
        if asset.data_center_short == failed_dc:
            asset.latest_operational_state = "OFFLINE"
            asset.write_authority = False
            # Demote roles if database
            if asset.latest_replication_role in ["PRIMARY", "PHYSICAL_STANDBY"]:
                asset.latest_replication_role = "SECONDARY" if asset.tech_stack == "mongodb" else "PHYSICAL_STANDBY"
            mutated_count += 1

        # If the asset resides in the promoted DC, make it active and promote it
        elif asset.data_center_short == promoted_dc:
            asset.latest_operational_state = "ACTIVE"
            # Promote standby roles
            if asset.latest_replication_role in ["SECONDARY", "PHYSICAL_STANDBY"]:
                asset.latest_replication_role = "PRIMARY"
                asset.write_authority = True
            mutated_count += 1

    # Log the event in Audit database
    audit = RuntimeAuditLog(
        id=str(uuid.uuid4()),
        event_type="FAILOVER_EXECUTED",
        description=f"Executed simulated failover for {app_id} in {environment}. Failed DC: {failed_dc}, Promoted DC: {promoted_dc}.",
        application_id=app_id,
        actor="operator"
    )
    db.add(audit)

    # Re-run drift detection
    drifts = await run_drift_detection(db, app_id, environment, persist_critical=True)
    alignment = compute_alignment_status(drifts)
    await db.execute(
        update(ApplicationIntent)
        .where(ApplicationIntent.application_id == app_id)
        .values(alignment_status=alignment)
    )

    await db.commit()

    return {
        "status": "SUCCESS",
        "message": f"Successfully executed simulated failover for {app_id}.",
        "mutated_assets_count": mutated_count,
        "alignment_status": alignment
    }


@router.post("/failback", response_model=Dict[str, Any])
async def execute_failback(
    data: Dict[str, Any],
    db: AsyncSession = Depends(get_db),
):
    """
    POST /api/v1/runtime-location/failback
    Body: { "application_id": "PCA", "environment": "PRODUCTION" }
    """
    app_id = data.get("application_id")
    environment = data.get("environment", "PRODUCTION")

    if not app_id:
        raise HTTPException(status_code=400, detail="Missing required field: application_id")

    # Fetch all assets for this app and environment
    result = await db.execute(select(RuntimeAsset))
    all_assets = result.scalars().all()

    app_assets = []
    for a in all_assets:
        if a.environment == environment:
            if a.metadata_json and a.metadata_json.get("application_id") == app_id:
                app_assets.append(a)
            elif app_id == "MQ_INFRA" and a.data_source == "ibm_mq":
                app_assets.append(a)
            elif app_id == "MONGO_INFRA" and a.data_source == "mongodb":
                app_assets.append(a)
            elif app_id == "ORACLE_INFRA" and a.data_source == "oracle_oem":
                app_assets.append(a)

    if not app_assets:
        raise HTTPException(status_code=404, detail="No assets found for the specified application and environment")

    # Revert to seeded defaults based on intent if it exists, or toggle
    intent_res = await db.execute(select(ApplicationIntent).where(ApplicationIntent.application_id == app_id))
    intent = intent_res.scalar_one_or_none()
    
    primary_dc = intent.intended_primary_dc if intent else "IBB1"

    mutated_count = 0
    for asset in app_assets:
        is_in_primary_dc = asset.data_center_short == primary_dc
        
        # Check tech stack specific defaults
        if asset.tech_stack == "oracle":
            if is_in_primary_dc:
                asset.latest_operational_state = "ACTIVE"
                asset.latest_replication_role = "PRIMARY"
                asset.write_authority = True
            else:
                asset.latest_operational_state = "STANDBY"
                asset.latest_replication_role = "PHYSICAL_STANDBY"
                asset.write_authority = False
        elif asset.tech_stack == "mongodb":
            if is_in_primary_dc:
                asset.latest_operational_state = "ACTIVE"
                asset.latest_replication_role = "PRIMARY"
                asset.write_authority = True
            else:
                asset.latest_operational_state = "STANDBY"
                asset.latest_replication_role = "SECONDARY"
                asset.write_authority = False
        elif asset.tech_stack == "mssql":
            if is_in_primary_dc:
                asset.latest_operational_state = "ACTIVE"
                asset.latest_replication_role = "PRIMARY"
                asset.write_authority = True
            else:
                asset.latest_operational_state = "STANDBY"
                asset.latest_replication_role = "SECONDARY"
                asset.write_authority = False
        elif asset.tech_stack == "ibm_mq":
            is_ga = asset.data_center_short == "GA-PRD"
            asset.latest_operational_state = "ACTIVE" if is_ga else "STANDBY"
            asset.write_authority = is_ga
        else:
            asset.latest_operational_state = "ACTIVE"
            asset.write_authority = False

        mutated_count += 1

    # Log the event in Audit database
    audit = RuntimeAuditLog(
        id=str(uuid.uuid4()),
        event_type="FAILBACK_EXECUTED",
        description=f"Executed simulated failback for {app_id} in {environment}. Restored primary DC: {primary_dc}.",
        application_id=app_id,
        actor="operator"
    )
    db.add(audit)

    # Re-run drift detection
    drifts = await run_drift_detection(db, app_id, environment, persist_critical=True)
    alignment = compute_alignment_status(drifts)
    await db.execute(
        update(ApplicationIntent)
        .where(ApplicationIntent.application_id == app_id)
        .values(alignment_status=alignment)
    )

    await db.commit()

    return {
        "status": "SUCCESS",
        "message": f"Successfully executed failback for {app_id} to primary DC {primary_dc}.",
        "mutated_assets_count": mutated_count,
        "alignment_status": alignment
    }


# ─── Snapshots ────────────────────────────────────────────────────────────────

@router.get("/snapshots/{app_id}", response_model=List[Dict[str, Any]])
async def get_snapshots(
    app_id: str,
    environment: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """
    GET /api/v1/runtime-location/snapshots/{app_id}
    Returns synthesized snapshot history from the current DB asset state.
    Each asset becomes a snapshot record representing its last-known state.
    Provides real persistence: data survives page refresh.
    """
    env = environment or "PRODUCTION"

    result = await db.execute(select(RuntimeAsset))
    all_assets = result.scalars().all()

    # Filter to assets matching this application
    app_assets = []
    for a in all_assets:
        if a.environment != env:
            continue
        if a.metadata_json and a.metadata_json.get("application_id") == app_id:
            app_assets.append(a)
        elif app_id == "MQ_INFRA" and a.data_source == "ibm_mq":
            app_assets.append(a)
        elif app_id == "MONGO_INFRA" and a.data_source == "mongodb":
            app_assets.append(a)
        elif app_id == "ORACLE_INFRA" and a.data_source == "oracle_oem":
            app_assets.append(a)

    snapshots = []
    for asset in app_assets:
        snapshots.append({
            "id": f"snap-{asset.id}",
            "asset_id": asset.id,
            "snapshot_time": asset.last_seen_at.isoformat() + "Z" if asset.last_seen_at else datetime.utcnow().isoformat() + "Z",
            "operational_state": asset.latest_operational_state or "UNKNOWN",
            "replication_role": asset.latest_replication_role or "NONE",
            "data_source": asset.data_source,
            "confidence_level": asset.latest_confidence_level or 3,
            "is_deterministic": asset.is_deterministic or False,
        })

    # Sort most-recent first
    snapshots.sort(key=lambda x: x["snapshot_time"], reverse=True)
    return snapshots


# ─── Cross-Environment Comparison ─────────────────────────────────────────────

@router.get("/compare-envs/{app_id}", response_model=List[Dict[str, Any]])
async def compare_environments(
    app_id: str,
    db: AsyncSession = Depends(get_db),
):
    """
    GET /api/v1/runtime-location/compare-envs/{app_id}
    Compares the same application across PRODUCTION, UAT, and DR environments.
    Returns a list of comparison rows showing role, DC, and confidence per env.
    """
    result = await db.execute(select(RuntimeAsset))
    all_assets = result.scalars().all()

    # Filter to assets matching this application across ALL environments
    app_assets = []
    for a in all_assets:
        is_match = False
        if a.metadata_json and a.metadata_json.get("application_id") == app_id:
            is_match = True
        elif app_id == "MQ_INFRA" and a.data_source == "ibm_mq":
            is_match = True
        elif app_id == "MONGO_INFRA" and a.data_source == "mongodb":
            is_match = True
        elif app_id == "ORACLE_INFRA" and a.data_source == "oracle_oem":
            is_match = True
        elif app_id == "INFRASTRUCTURE" and not (a.metadata_json and a.metadata_json.get("application_id")):
            is_match = True
        if is_match:
            app_assets.append(a)

    if not app_assets:
        return []

    # Group assets by a normalized name key to match across environments
    # Normalize: strip environment suffixes, lowercase
    def normalize_asset_name(asset):
        name = asset.name.lower().strip()
        # Remove common env suffixes for grouping
        for suffix in ["_primary", "_standby", "-primary", "-standby"]:
            if name.endswith(suffix):
                name = name[: -len(suffix)]
        return name

    name_groups = {}
    for a in app_assets:
        key = normalize_asset_name(a)
        if key not in name_groups:
            name_groups[key] = {}
        env = a.environment
        # Store the best (highest confidence) asset per env
        if env not in name_groups[key] or a.latest_confidence_level > name_groups[key][env].latest_confidence_level:
            name_groups[key][env] = a

    comparison_rows = []
    for norm_name, env_map in name_groups.items():
        # Pick a representative asset for display name / tech_stack / component
        rep = next(iter(env_map.values()))
        comp_name = f"{rep.tech_stack.replace('_', ' ').upper()} Layer"

        prod = env_map.get("PRODUCTION")
        uat = env_map.get("UAT")
        dr = env_map.get("DR")

        # Determine consistency status
        present_envs = [e for e in ["PRODUCTION", "UAT", "DR"] if e in env_map]
        roles = list(set(
            env_map[e].latest_replication_role
            for e in present_envs
            if env_map[e].latest_replication_role
        ))

        if len(present_envs) == 1:
            env_key = present_envs[0]
            status = "prod_only" if env_key == "PRODUCTION" else "uat_only" if env_key == "UAT" else "dr_only"
        elif len(roles) <= 1:
            status = "consistent"
        else:
            # Different roles across environments is expected (PRIMARY in prod, STANDBY in DR)
            # Only flag as inconsistent if same-role expectations conflict
            has_multi_primary = sum(
                1 for e in present_envs
                if env_map[e].latest_replication_role in ["PRIMARY", "ACTIVE"]
            ) > 1
            status = "inconsistent" if has_multi_primary else "consistent"

        row = {
            "asset_name": rep.name,
            "tech_stack": rep.tech_stack,
            "component": comp_name,
            "prod_role": prod.latest_replication_role if prod else None,
            "prod_dc": prod.data_center_short if prod else None,
            "prod_confidence": prod.latest_confidence_level if prod else None,
            "uat_role": uat.latest_replication_role if uat else None,
            "uat_dc": uat.data_center_short if uat else None,
            "uat_confidence": uat.latest_confidence_level if uat else None,
            "dr_role": dr.latest_replication_role if dr else None,
            "dr_dc": dr.data_center_short if dr else None,
            "dr_confidence": dr.latest_confidence_level if dr else None,
            "status": status,
        }
        comparison_rows.append(row)

    # Sort: inconsistent first, then by asset name
    status_order = {"inconsistent": 0, "prod_only": 1, "uat_only": 2, "dr_only": 3, "consistent": 4}
    comparison_rows.sort(key=lambda r: (status_order.get(r["status"], 5), r["asset_name"]))

    return comparison_rows
